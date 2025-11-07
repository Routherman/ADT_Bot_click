#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Push and sync utility for the general sheet 'Todos_ADT_OTS_Expo'.

Tabs:
  1) procesados_bot
     Columns: Item_Valido,comentario,ID,Nombre,Dirección,Teléfono,Sitio Web,Ciudad,procesado_Manual,Venue,puntaje_web,emails_Web,emails_externos,CantidadEmails,social1..6,FechaExport
  2) sin_procesar_bot
     Columns: ID, Venue, Nombre, Dirección, Teléfono, Sitio Web, Ciudad, GoogleMapsURL
  3) procesados_paratanda2 (descartados)
     Same columns as (1)

Real-time usage: call push_site(site_obj) after nav_pro processing. It will dedupe by domain and decide target tab
according to config_map filters, score threshold, and Florida zone.
Manual usage: call rebuild_unprocessed_from_csv() for tab (2), and sync_processed_from_etapa1() for tabs (1) and (3).

Requires: gspread + service account (GOOGLE_APPLICATION_CREDENTIALS).
"""
from __future__ import annotations
import os, json, re, time
from typing import Any, Dict, List, Optional, Tuple, Set
from datetime import datetime, timezone
from pathlib import Path
import sqlite3

# Defaults
DEFAULT_SHEET_ID = os.getenv('TODOS_EXPO_SHEET_ID', '1TpmmZoRThgMqIeXhN3d14klEkMhlA1UvP6tXvAd1Qh8').strip()
TAB1 = 'procesados_bot'
TAB2 = 'sin_procesar_bot'
TAB3 = 'procesados_paratanda2'
TAB4 = 'Filtrados_sin_procesar_bot'

TAB1_HEADER = [
    'Item_Valido','comentario','ID','Nombre','Dirección','Teléfono','Sitio Web','Ciudad',
    'procesado_Manual','Venue','puntaje_web','emails_Web','emails_externos','CantidadEmails',
    'social1','social2','social3','social4','social5','social6','FechaExport','Short_descrip'
]
TAB2_HEADER = ['ID','Venue','Nombre','Dirección','Teléfono','Sitio Web','Ciudad','GoogleMapsURL']
TAB3_HEADER = [
    'Item_Valido','comentario','motivo_descarte','ID','Nombre','Dirección','Teléfono','Sitio Web','Ciudad',
    'procesado_Manual','Venue','puntaje_web','emails_Web','emails_externos','CantidadEmails',
    'social1','social2','social3','social4','social5','social6','FechaExport'
]

# Nuevo TAB4: Filtrados por DOM_Name (descartados antes de procesar)
TAB4_HEADER = ['ID','Venue','Nombre','Dirección','Teléfono','Sitio Web','Ciudad','GoogleMapsURL','Filtro_hit']

_EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@(?:[A-Z0-9\-]+\.)+[A-Z]{2,24}", re.I)

_gc = None
_ws_cache: Dict[str, Any] = {}
_present_domains: Dict[str, Set[str]] = {TAB1: set(), TAB3: set()}
_initialized = False
_coment_idx: Dict[tuple, Dict[str, Any]] = {}
_csv_index: Dict[str, Dict[str, Any]] = {}
_excl_cache: Dict[str, Any] = {}

# ---------- Inclusion (mark-as-SI) sheet helpers ----------
_incl_cache: Dict[str, Any] = {}

def _load_mark_si_sets() -> Tuple[Set[str], Set[str]]:
    """Loads sets (domains_set, emails_set) from a Google Sheet to mark items as SI.
    Sheet ID priority: TAB1_MARK_SI_SHEET_ID env > ENRICH_EXCLUDE_SHEET_ID env (fallback).
    Uses exportar_etapa1.load_usa_exclusion_sets for convenience (returns names, emails, domains, web_domains).
    Caches by (sheet_id|cred).
    """
    try:
        sid = os.getenv('TAB1_MARK_SI_SHEET_ID', '').strip() or os.getenv('ENRICH_EXCLUDE_SHEET_ID','').strip()
        if not sid:
            return set(), set()
        cred = os.getenv('GOOGLE_APPLICATION_CREDENTIALS','').strip() or ''
        key = f"{sid}|{cred}"
        if key in _incl_cache:
            return _incl_cache[key]
        try:
            from exportar_etapa1 import load_usa_exclusion_sets  # type: ignore
        except Exception:
            return set(), set()
        data = load_usa_exclusion_sets(sid, cred or None)
        domains = set((data.get('domains') or set()) | (data.get('web_domains') or set()))
        emails = set(data.get('emails') or set())
        _incl_cache[key] = (domains, emails)
        return _incl_cache[key]
    except Exception:
        return set(), set()

def _site_emails_simple(site_obj: Dict[str, Any]) -> Set[str]:
    out: Set[str] = set()
    try:
        for e in (site_obj.get('emails') or []):
            if isinstance(e, dict):
                v = (e.get('value') or e.get('email') or e.get('mail') or '').strip().lower()
            else:
                v = str(e).strip().lower()
            if v:
                out.add(v)
    except Exception:
        pass
    return out

def _is_marked_yes(site_obj: Dict[str, Any]) -> bool:
    try:
        doms, mails = _load_mark_si_sets()
        if not (doms or mails):
            return False
        d = (site_obj.get('domain') or '').strip().lower().lstrip('www.')
        if d and d in doms:
            return True
        for e in _site_emails_simple(site_obj):
            if e in mails:
                return True
        return False
    except Exception:
        return False

# Feature flags (dynamic stop for exports)
FEATURE_FLAGS_FILE = os.getenv('FEATURE_FLAGS_FILE', os.path.join('out','feature_flags.json'))

def _exports_disabled() -> bool:
    # env override
    env_block = os.getenv('EXPORTS_DISABLED')
    if env_block and env_block.strip().lower() in ('1','true','yes','on'):  # allow dynamic via env too
        return True
    try:
        if os.path.exists(FEATURE_FLAGS_FILE):
            flags = json.load(open(FEATURE_FLAGS_FILE,'r',encoding='utf-8'))
            return bool(flags.get('exports_disabled'))
    except Exception:
        return False
    return False


def _domain_from_url(web: str) -> str:
    try:
        from urllib.parse import urlparse
        p = urlparse((web or '').strip())
        d = (p.netloc or '').lower()
        return d.lstrip('www.')
    except Exception:
        return (web or '').strip().lower().lstrip('www.')


def _extract_emails_values(values: List[Any]) -> List[str]:
    out: List[str] = []
    for v in (values or []):
        raw = (v.get('value') if isinstance(v, dict) else str(v)) or ''
        raw = str(raw)
        for m in _EMAIL_RE.finditer(raw):
            out.append(m.group(0).lower())
    # dedupe preserve order
    seen=set(); res=[]
    for e in out:
        if e not in seen:
            seen.add(e); res.append(e)
    return res

def _first_phone(site_obj: Dict[str,Any], row: Dict[str,Any]) -> str:
    # CSV columns first
    tel = row.get('Teléfono', row.get('Telefono', row.get('Phone', row.get('TEL',''))))
    if tel:
        return str(tel)
    # Try structured phones list
    try:
        for p in site_obj.get('phones') or []:
            if isinstance(p, dict) and p.get('value'):
                return str(p.get('value'))
            if isinstance(p, str) and p.strip():
                return p.strip()
    except Exception:
        pass
    return ''

def _infer_city(site_obj: Dict[str,Any], row: Dict[str,Any]) -> str:
    c = row.get('Ciudad', row.get('City',''))
    if c:
        return str(c)
    try:
        addrs = site_obj.get('addresses') or []
        if addrs and isinstance(addrs[0], dict):
            for key in ('city','locality','town','municipality'):
                v = addrs[0].get(key)
                if v:
                    return str(v)
    except Exception:
        pass
    return ''

def _site_web(site_obj: Dict[str,Any], row: Dict[str,Any], dom: str) -> str:
    w = row.get('Sitio Web', row.get('Website', row.get('URL','')))
    if w:
        return str(w)
    if site_obj.get('site_url'):
        return str(site_obj.get('site_url'))
    if dom:
        return f"https://{dom}"
    return ''

def _site_name(site_obj: Dict[str,Any], row: Dict[str,Any], dom: str) -> str:
    n = row.get('Nombre', row.get('Name',''))
    if n:
        return str(n)
    for k in ('site_name','name','nombre'):
        v = site_obj.get(k)
        if v:
            return str(v)
    # fallback from domain
    if dom:
        base = dom.split('.')[0]
        return base.replace('-', ' ').title()
    return ''

def _resolve_id(site_obj: Dict[str,Any], row: Dict[str,Any]) -> str:
    _id = row.get('ID', row.get('Id', row.get('id','')))
    if _id:
        return str(_id)
    if site_obj.get('id'):
        return str(site_obj.get('id'))
    return ''

def _infer_venue(src: Dict[str,Any], row: Dict[str,Any]) -> str:
    r = src.get('rubro')
    if r:
        return str(r)
    v = row.get('Venue')
    if v:
        return str(v)
    filep = src.get('file') or src.get('csv_path') or src.get('csv_name')
    try:
        if filep:
            return Path(str(filep)).stem
    except Exception:
        pass
    return ''

# ------------- Local backup (xlsx preferred, csv fallback) -------------
_backup_dir = os.path.join('out','backup')

def _append_local_backup_rows(tab: str, rows: List[List[str]]):
    os.makedirs(_backup_dir, exist_ok=True)
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
        xlsx_path = os.path.join(_backup_dir, 'Todos_local.xlsx')
        if os.path.exists(xlsx_path):
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()
            ws0 = wb.active; ws0.title = TAB1
            for t in (TAB2, TAB3):
                wb.create_sheet(title=t)
            wb[TAB1].append(TAB1_HEADER)
            wb[TAB2].append(TAB2_HEADER)
            wb[TAB3].append(TAB3_HEADER)
        if tab not in wb.sheetnames:
            wb.create_sheet(title=tab)
        ws = wb[tab]
        # If empty, seed header
        if ws.max_row == 1 and (ws.cell(1,1).value is None):
            if tab == TAB1: ws.append(TAB1_HEADER)
            elif tab == TAB2: ws.append(TAB2_HEADER)
            else: ws.append(TAB3_HEADER)
        for r in rows:
            ws.append(r)
        wb.save(xlsx_path)
        return
    except Exception:
        pass
    # Fallback to CSV append
    import csv as _csv
    csv_path = os.path.join(_backup_dir, f'{tab}.csv')
    need_header = not os.path.exists(csv_path)
    try:
        with open(csv_path, 'a', newline='', encoding='utf-8') as f:
            w = _csv.writer(f)
            if need_header:
                if tab == TAB1: w.writerow(TAB1_HEADER)
                elif tab == TAB2: w.writerow(TAB2_HEADER)
                else: w.writerow(TAB3_HEADER)
            for r in rows:
                w.writerow(r)
    except Exception:
        pass

def _socials_order(site_obj: Dict[str, Any]) -> List[str]:
    try:
        from social_utils import prioritize_socials  # type: ignore
    except Exception:
        prioritize_socials = lambda seq, base_limit=6, allow_overflow_priorities=False: list(seq)  # type: ignore
    raw = []
    for s in (site_obj.get('socials') or []):
        if isinstance(s, dict):
            u = (s.get('url') or '').strip()
        else:
            u = str(s).strip()
        if u:
            raw.append(u)
    ordered = prioritize_socials(raw, base_limit=6, allow_overflow_priorities=False)
    return (ordered + ['']*6)[:6]


def _load_sheet(sheet_id: Optional[str]=None):
    global _gc
    import gspread  # type: ignore
    sid = (sheet_id or DEFAULT_SHEET_ID)
    cred = os.getenv('GOOGLE_APPLICATION_CREDENTIALS','').strip() or None
    if _gc is None:
        if cred and os.path.isfile(cred):
            _gc = gspread.service_account(filename=cred)
        else:
            _gc = gspread.service_account()
    return _gc.open_by_key(sid)


def _apply_validations_for_ws(sh, ws, title: str) -> None:
    """Apply dropdown validations for Item_Valido (A) and comentario (B). Best-effort."""
    try:
        sheet_id = ws._properties.get('sheetId')
        if not sheet_id:
            return
        # Item_Valido dropdown exact values/order requested
        # Note: must match the value we write by default elsewhere ('NO_Revisado')
        item_values = ['NO_Revisado','SI','NO']
        comment_values = [
            'NO hace eventos de musica',
            'Espacio reducido',
            'No contratria banda',
            'Listador de eventos',
            'Venta de Tickects o Agencia',
            'OTRO'
        ]
        def _val_list(vals: List[str]):
            return [{"userEnteredValue": v} for v in vals]
        requests = [
            {
                "setDataValidation": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 20000, "startColumnIndex": 0, "endColumnIndex": 1},
                    "rule": {"condition": {"type": "ONE_OF_LIST", "values": _val_list(item_values)}, "strict": True, "showCustomUi": True}
                }
            },
            {
                "setDataValidation": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 20000, "startColumnIndex": 1, "endColumnIndex": 2},
                    "rule": {"condition": {"type": "ONE_OF_LIST", "values": _val_list(comment_values)}, "strict": False, "showCustomUi": True}
                }
            }
        ]
        sh.batch_update({"requests": requests})
    except Exception:
        pass


def _ensure_ws(sh, title: str, header: List[str]):
    try:
        ws = None
        try:
            ws = sh.worksheet(title)
        except Exception:
            ws = sh.add_worksheet(title=title, rows=2000, cols=max(20, len(header)))
        values = ws.get_all_values() or []
        if not values:
            ws.update([header])
        else:
            # If header mismatch, rewrite header row
            if values[0] != header:
                ws.update([header])
        if title in (TAB1, TAB3):
            _apply_validations_for_ws(sh, ws, title)
        return ws
    except Exception:
        return None


def _ensure_initialized():
    global _initialized
    if _initialized:
        return
    sh = _load_sheet()
    ws1 = _ensure_ws(sh, TAB1, TAB1_HEADER); ws2 = _ensure_ws(sh, TAB2, TAB2_HEADER); ws3 = _ensure_ws(sh, TAB3, TAB3_HEADER)
    ws4 = _ensure_ws(sh, TAB4, TAB4_HEADER)
    if ws1: _ws_cache[TAB1] = ws1
    if ws2: _ws_cache[TAB2] = ws2
    if ws3: _ws_cache[TAB3] = ws3
    if ws4: _ws_cache[TAB4] = ws4
    # Seed present domains for TAB1 and TAB3 dedupe from current sheet values (use 'Sitio Web' -> domain or fallback from "Dominio" not present here)
    try:
        for tab in (TAB1, TAB3):
            ws = _ws_cache.get(tab)
            if not ws: continue
            vals = ws.get_all_values() or []
            if not vals or len(vals) < 2: continue
            header = vals[0]
            try:
                idx_web = header.index('Sitio Web')
            except ValueError:
                idx_web = -1
            doms: Set[str] = set()
            for row in vals[1:]:
                web = (row[idx_web] if (idx_web >= 0 and idx_web < len(row)) else '').strip()
                if web:
                    d = _domain_from_url(web)
                    if d:
                        doms.add(d)
            _present_domains[tab] = doms
    except Exception:
        pass
    _initialized = True


def update_stats_tab(sheet_id: Optional[str]=None) -> Dict[str, Any]:
    """Compute per-rubro (Venue) metrics from TAB1 ('procesados_bot') and write snapshot to 'stats' tab.
    Columns written:
      timestamp, rubro, total_items, validos, pct_validos, pct_no_revisado, avg_cantidad_emails, median_cantidad_emails
    Rules:
      - 'validos' counts Item_Valido in {SI, HECHO} (case-insensitive)
      - pct_validos = (validos / total_items) * 100
      - pct_no_revisado = (NO_Revisado / total_items) * 100
      - avg_cantidad_emails uses numeric values from 'CantidadEmails' (non-numeric ignored)
      - median_cantidad_emails computed on numeric values only
    Returns a small summary dict with totals and number of rubros summarized.
    """
    try:
        _ensure_initialized()
        ws1 = _ws_cache.get(TAB1)
        if not ws1:
            return {"ok": False, "error": "TAB1 not available"}
        values = ws1.get_all_values() or []
        if not values or len(values) < 2:
            # create empty stats tab with header
            sh = _load_sheet(sheet_id)
            ws_stats = _ensure_ws(sh, 'stats', ['timestamp','rubro','total_items','validos','pct_validos','pct_no_revisado','avg_cantidad_emails','median_cantidad_emails'])
            if ws_stats:
                try:
                    ws_stats.clear()
                    ws_stats.update('A1', [['timestamp','rubro','total_items','validos','pct_validos','pct_no_revisado','avg_cantidad_emails','median_cantidad_emails']])
                except Exception:
                    pass
            return {"ok": True, "rubros": 0, "total_items": 0}
        header = values[0]
        def _hidx(name: str) -> int:
            try:
                return header.index(name)
            except ValueError:
                return -1
        idx_item = _hidx('Item_Valido'); idx_venue = _hidx('Venue'); idx_cant = _hidx('CantidadEmails')
        if idx_item < 0 or idx_venue < 0:
            return {"ok": False, "error": "Required columns not found in TAB1"}
        from collections import defaultdict
        agg = defaultdict(lambda: {"total":0, "valid":0, "no_rev":0, "sum_em":0, "em_count_n":0, "em_vals": []})
        total_items = 0
        for row in values[1:]:
            try:
                rub = (row[idx_venue] if idx_venue < len(row) else '').strip()
                item = (row[idx_item] if idx_item < len(row) else '').strip().upper()
                val_em = (row[idx_cant] if (idx_cant >= 0 and idx_cant < len(row)) else '').strip()
                em = None
                if val_em:
                    try:
                        em = int(float(val_em))
                    except Exception:
                        try:
                            em = int(val_em)
                        except Exception:
                            em = None
                bucket = agg[rub]
                bucket['total'] += 1; total_items += 1
                if item in ('SI','HECHO'): bucket['valid'] += 1
                if item == 'NO_REVISADO': bucket['no_rev'] += 1
                if em is not None:
                    em = max(0, em)
                    bucket['sum_em'] += em
                    bucket['em_count_n'] += 1
                    bucket['em_vals'].append(em)
            except Exception:
                continue
        from datetime import datetime, timezone
        ts = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00','Z')
        rows_out: List[List[str]] = [[
            'timestamp','rubro','total_items','validos','pct_validos','pct_no_revisado','avg_cantidad_emails','median_cantidad_emails'
        ]]
        rubros = 0
        for rubro, m in sorted(agg.items(), key=lambda kv: (kv[0] or '').lower()):
            rubros += 1
            tot = m['total'] or 0
            pct_valid = (m['valid'] / tot * 100.0) if tot else 0.0
            pct_no = (m['no_rev'] / tot * 100.0) if tot else 0.0
            avg_em = (m['sum_em'] / m['em_count_n']) if m['em_count_n'] else 0.0
            try:
                import statistics as _stats
                med_em = float(_stats.median(m['em_vals'])) if m['em_vals'] else 0.0
            except Exception:
                med_em = 0.0
            rows_out.append([
                ts, rubro or '', str(tot), str(m['valid']), f"{pct_valid:.2f}", f"{pct_no:.2f}", f"{avg_em:.2f}", f"{med_em:.2f}"
            ])
        sh = _load_sheet(sheet_id)
        ws_stats = _ensure_ws(sh, 'stats', rows_out[0])
        if ws_stats:
            try: ws_stats.clear()
            except Exception: pass
            try: ws_stats.update('A1', rows_out)
            except Exception:
                try: ws_stats.append_rows(rows_out, value_input_option='RAW')
                except Exception: pass
        return {"ok": True, "rubros": rubros, "total_items": total_items}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------- Config filters (mirrors exportar/nav_pro minimal subset) ----------

def _normalize_rubro_key(rubro: Optional[str]) -> str:
    rl = (rubro or '').strip().lower()
    if rl.startswith('rubro_'):
        rl = rl[6:]
    return rl


def _kw_hits_dom_or_name(domain: str, site_name: str, kws: List[str]) -> int:
    if not kws: return 0
    d = (domain or '').lower(); n = (site_name or '').lower(); tot = 0
    for k in kws:
        kk = str(k).strip().lower()
        if kk and (kk in d or kk in n):
            tot += 1
    return tot


def _kw_hits_in_item(site: Dict[str, Any], kws: List[str]) -> int:
    if not kws: return 0
    tot = 0
    dom = str(site.get('domain') or '').lower(); surl = str(site.get('site_url') or '').lower()
    band = site.get('band') if isinstance(site, dict) else {}
    per = band.get('per_keyword') if isinstance(band, dict) else {}
    per_keys_low = [str(x).lower() for x in (per.keys() if isinstance(per, dict) else [])]
    for k in kws:
        kk = str(k).strip().lower()
        if not kk: continue
        if (kk in dom) or (kk in surl) or any(kk in pk for pk in per_keys_low):
            tot += 1
    return tot


# ---------- Short_descrip backfill ----------
def backfill_short_descrip_tab1(sheet_id: Optional[str]=None, *, only_empty: bool=True, limit: Optional[int]=None) -> int:
    """Fill Short_descrip for rows in TAB1 using AI generator with cache (no network required if no API).
    Returns number of rows updated.
    """
    _ensure_initialized()
    ws1 = _ws_cache.get(TAB1)
    if not ws1:
        return 0
    vals = ws1.get_all_values() or []
    if not vals:
        return 0
    header = vals[0]
    def _idx(col: str) -> int:
        try:
            return header.index(col)
        except ValueError:
            return -1
    idx_web = _idx('Sitio Web')
    idx_city = _idx('Ciudad')
    idx_venue = _idx('Venue')
    idx_score = _idx('puntaje_web')
    idx_short = _idx('Short_descrip')
    if idx_web < 0 or idx_short < 0:
        return 0
    updates = []  # list of (row_idx, value)
    total_rows = len(vals) - 1
    count = 0
    from ai_short_descrip import generate_short_description  # type: ignore
    for i in range(1, len(vals)):
        if limit is not None and count >= limit:
            break
        row = vals[i]
        cur_short = (row[idx_short] if idx_short < len(row) else '').strip()
        if only_empty and cur_short:
            continue
        web = (row[idx_web] if idx_web < len(row) else '').strip()
        if not web:
            continue
        # derive domain
        dom = _domain_from_url(web)
        if not dom:
            continue
        # Build minimal site context
        city = (row[idx_city] if idx_city >= 0 and idx_city < len(row) else '').strip()
        venue = (row[idx_venue] if idx_venue >= 0 and idx_venue < len(row) else '').strip()
        score = (row[idx_score] if idx_score >= 0 and idx_score < len(row) else '').strip()
        site_ctx = {
            'domain': dom,
            'addresses': [{'city': city}] if city else [],
            'source_csv': {'rubro': venue} if venue else {},
            'band': {'score': int(score) if str(score).isdigit() else None}
        }
        try:
            sdesc = generate_short_description(site_ctx, cache_dir=os.path.join('out','etapa1_cache'))
        except Exception:
            sdesc = ''
        sdesc = (sdesc or '').strip()
        if not sdesc:
            continue
        updates.append((i+1, sdesc))  # 1-based row index on sheet
        count += 1
    if not updates:
        return 0
    # Batch update Short_descrip column efficiently
    # Compute A1 ranges for each row in Short_descrip col
    def col_letter(idx_zero: int) -> str:
        # Convert zero-based index to A1 column letter
        idx = idx_zero + 1
        letters = ''
        while idx:
            idx, rem = divmod(idx - 1, 26)
            letters = chr(65 + rem) + letters
        return letters
    col = col_letter(idx_short)
    # chunk updates to avoid payload limits
    chunks = [updates[i:i+200] for i in range(0, len(updates), 200)]
    for ch in chunks:
        reqs = []
        for r, val in ch:
            rng = f"{col}{r}:{col}{r}"
            reqs.append({'range': rng, 'values': [[val]]})
        try:
            ws1.batch_update(reqs)
        except Exception:
            # fallback cell by cell
            for r, val in ch:
                try:
                    ws1.update(f"{col}{r}", val)
                except Exception:
                    pass
        time.sleep(0.2)
    return len(updates)


def _load_config_map() -> Dict[str, Any]:
    path = os.getenv('CONFIG_MAP_PATH', os.path.join('configuraciones','config_map.json'))
    try:
        return json.load(open(path,'r',encoding='utf-8'))
    except Exception:
        return {}


def _evaluate_filters(site: Dict[str, Any]) -> Tuple[bool, str, float]:
    """Return (discarded, reason, effective_score_for_threshold).
    Applies config_map discard (defaults + rubro + dom_name), boosts/penalties (light), and basic threshold.
    Also considers Florida invalidation via site['florida_ok'].
    """
    cfg = _load_config_map() or {}
    cfg_defaults = (cfg.get('defaults') if isinstance(cfg, dict) else {}) or {}
    cfg_defaults_dom = (cfg.get('defaults_Dom_Name') if isinstance(cfg, dict) else {}) or {}
    cfg_rubros = (cfg.get('rubros') if isinstance(cfg, dict) else {}) or {}
    cfg_scoring = (cfg.get('scoring') if isinstance(cfg, dict) else {}) or {}
    boost_delta = int((cfg_scoring.get('boost_delta') if isinstance(cfg_scoring, dict) else 8) or 8)
    penalty_delta = int((cfg_scoring.get('penalty_delta') if isinstance(cfg_scoring, dict) else 12) or 12)
    discard_min_hits = int((cfg_scoring.get('discard_min_hits') if isinstance(cfg_scoring, dict) else 1) or 1)
    # rubro key and source name
    src = site.get('source_csv') or {}; row = (src.get('row') or {}) if isinstance(src, dict) else {}
    name_src = ''
    try:
        if isinstance(row, dict):
            name_src = str(row.get('Nombre', row.get('Name','')) or '')
    except Exception:
        name_src = ''
    rubro = (src.get('rubro') or '') if isinstance(src, dict) else ''
    rkey = _normalize_rubro_key(rubro)
    r_cfg = (cfg_rubros.get(rkey) if isinstance(cfg_rubros, dict) else None) or {}
    # Discard keywords (content) + dom_name keywords (also consider Address text for dom_name)
    disc_kws = []
    if isinstance(cfg_defaults, dict):
        disc_kws += [str(x) for x in (cfg_defaults.get('discard_keywords') or [])]
    if isinstance(r_cfg, dict):
        disc_kws += [str(x) for x in (r_cfg.get('discard_keywords') or [])]
    domname_disc = []
    if isinstance(cfg_defaults_dom, dict):
        domname_disc += [str(x) for x in (cfg_defaults_dom.get('discard_keywords') or [])]
    if isinstance(r_cfg, dict) and isinstance(r_cfg.get('dom_name'), dict):
        domname_disc += [str(x) for x in (r_cfg.get('dom_name').get('discard_keywords') or [])]
    # Discard check
    hits_disc = _kw_hits_in_item(site, disc_kws)
    hits_domname = _kw_hits_dom_or_name(site.get('domain') or '', name_src, domname_disc)
    # Address-aware discard (for cases like "Church Ave." / "Church Rd.")
    addr_text = ''
    try:
        addrs = site.get('addresses') or []
        if addrs and isinstance(addrs[0], dict):
            addr_text = str(addrs[0].get('value') or '')
    except Exception:
        pass
    try:
        src = site.get('source_csv') or {}
        row = (src.get('row') or {}) if isinstance(src, dict) else {}
        if not addr_text:
            addr_text = str(row.get('Dirección', row.get('Address','')) or '')
    except Exception:
        pass
    addr_low = addr_text.strip().lower()
    hits_addr_domname = 0
    if addr_low and domname_disc:
        for k in domname_disc:
            kk = str(k).strip().lower()
            if kk and kk in addr_low:
                hits_addr_domname += 1
    if (hits_disc + hits_domname + hits_addr_domname) >= max(1, discard_min_hits):
        return (True, f'discard:{hits_disc+hits_domname+hits_addr_domname} (dom_name={hits_domname},addr={hits_addr_domname})', 0.0)
    # Effective score with light boost/penalty
    try:
        base = float(((site.get('band') or {}).get('score')) or site.get('band_score') or 0)
    except Exception:
        base = float(site.get('band_score') or 0) if site.get('band_score') is not None else 0.0
    eff = base
    # boosts/penalties (content)
    boost_k = []; pen_k = []
    if isinstance(cfg_defaults, dict):
        boost_k += [str(x) for x in (cfg_defaults.get('boost_keywords') or [])]
        pen_k += [str(x) for x in (cfg_defaults.get('penalty_keywords') or [])]
    if isinstance(r_cfg, dict):
        boost_k += [str(x) for x in (r_cfg.get('boost_keywords') or [])]
        pen_k += [str(x) for x in (r_cfg.get('penalty_keywords') or [])]
    b_hits = _kw_hits_in_item(site, boost_k); p_hits = _kw_hits_in_item(site, pen_k)
    if b_hits >= 1: eff = min(100.0, eff + boost_delta)
    if p_hits >= 1: eff = eff - penalty_delta  # permitir negativos
    # dom_name quick adjustments
    domname_boost = []; domname_penalty = []
    if isinstance(cfg_defaults_dom, dict):
        domname_boost += [str(x) for x in (cfg_defaults_dom.get('boost_keywords') or [])]
        domname_penalty += [str(x) for x in (cfg_defaults_dom.get('penalty_keywords') or [])]
    if isinstance(r_cfg, dict) and isinstance(r_cfg.get('dom_name'), dict):
        domname_boost += [str(x) for x in (r_cfg.get('dom_name').get('boost_keywords') or [])]
        domname_penalty += [str(x) for x in (r_cfg.get('dom_name').get('penalty_keywords') or [])]
    if _kw_hits_dom_or_name(site.get('domain') or '', name_src, domname_boost) >= 1:
        eff = min(100.0, eff + boost_delta)
    if _kw_hits_dom_or_name(site.get('domain') or '', name_src, domname_penalty) >= 1:
        eff = eff - penalty_delta  # permitir negativos
    return (False, '', eff)


# ---------- USA exclusion sheet helpers ----------

def _load_usa_exclusions() -> Tuple[Set[str], Set[str], Set[str], Set[str]]:
    """Returns (excl_domains, excl_emails, excl_names, excl_web_domains). Caches results per process.
    If ENRICH_EXCLUDE_SHEET_ID is missing or loading fails, returns empty sets.
    """
    try:
        # Cache key based on sheet id + cred path
        sid = os.getenv('ENRICH_EXCLUDE_SHEET_ID', '').strip()
        if not sid:
            return set(), set(), set(), set()
        cred = os.getenv('GOOGLE_APPLICATION_CREDENTIALS','').strip() or ''
        key = f"{sid}|{cred}"
        if key in _excl_cache:
            return _excl_cache[key]
        try:
            from exportar_etapa1 import _load_exclusions_cached, load_usa_exclusion_sets  # type: ignore
        except Exception:
            return set(), set(), set(), set()
        # Force refresh to avoid stale cache from prior runs
        prev_refresh = os.environ.get('EXCLUSIONS_CACHE_REFRESH')
        os.environ['EXCLUSIONS_CACHE_REFRESH'] = '1'
        excl_domains, excl_emails = _load_exclusions_cached(sid, cred or None)
        full = load_usa_exclusion_sets(sid, cred or None)
        excl_names = full.get('names', set()) or set()
        excl_web_domains = full.get('web_domains', set()) or set()
        # Restore flag
        if prev_refresh is None:
            os.environ.pop('EXCLUSIONS_CACHE_REFRESH', None)
        else:
            os.environ['EXCLUSIONS_CACHE_REFRESH'] = prev_refresh
        _excl_cache[key] = (set(excl_domains), set(excl_emails), set(excl_names), set(excl_web_domains))
        return _excl_cache[key]
    except Exception:
        return set(), set(), set(), set()


def _is_excluded_usa(site_obj: Dict[str, Any]) -> bool:
    """Use USA exclusion sheet to decide exclusion by domain, website, name, or emails.
    Always returns False if exclusions are not configured/available.
    """
    try:
        excl_domains, excl_emails, excl_names, excl_web_domains = _load_usa_exclusions()
        if not (excl_domains or excl_emails or excl_names or excl_web_domains):
            return False
        # Domain checks
        dom = str(site_obj.get('domain') or '').strip().lower().lstrip('www.')
        if dom and (dom in excl_domains or dom in excl_web_domains):
            return True
        # Name check (use best-effort name resolution)
        src = site_obj.get('source_csv') or {}
        row = (src.get('row') or {}) if isinstance(src, dict) else {}
        name = _site_name(site_obj, row, dom)
        if name and name.strip().lower() in excl_names:
            return True
        # Emails (web + external) quick check from site_obj.emails raw list
        raw_emails = _extract_emails_values(site_obj.get('emails') or [])
        if any((e.strip().lower() in excl_emails) for e in raw_emails if e.strip()):
            return True
        return False
    except Exception:
        return False


# ---------- CSV index helpers (to fill ID/Venue when missing) ----------

def _ensure_csv_index(csv_dir: str = 'csv') -> None:
    """Build a domain -> {file, rubro, row} index from csv/*.csv once."""
    global _csv_index
    if _csv_index:
        return
    from pathlib import Path as _P
    import csv as _csv
    base = _P(csv_dir)
    if not base.exists():
        return
    for p in sorted(base.glob('*.csv')):
        rubro = p.stem
        try:
            with p.open('r', encoding='utf-8', errors='ignore') as f:
                reader = _csv.DictReader(f)
                for r in reader:
                    web = r.get('Sitio Web') or r.get('Website') or r.get('URL') or ''
                    dom = _domain_from_url(web) if web else ''
                    if not dom:
                        continue
                    if dom not in _csv_index:
                        _csv_index[dom] = {'file': str(p), 'rubro': rubro, 'row': r}
        except Exception:
            continue


def _attach_source_from_csv_if_missing(site_obj: Dict[str, Any]) -> None:
    """If site has no source_csv or lacks row/venue, try to enrich from CSV index by domain."""
    try:
        src = site_obj.get('source_csv') or {}
        has_row = isinstance(src, dict) and isinstance(src.get('row'), dict) and bool(src.get('row'))
        has_rubro = isinstance(src, dict) and bool(src.get('rubro'))
        if has_row and has_rubro:
            return
        dom = str(site_obj.get('domain') or '').strip().lower().lstrip('www.')
        if not dom:
            return
        _ensure_csv_index()
        hit = _csv_index.get(dom)
        if not hit:
            return
        site_obj['source_csv'] = {
            'file': hit.get('file'),
            'rubro': hit.get('rubro'),
            'row': hit.get('row') or {}
        }
    except Exception:
        pass


def _format_tab1_row(site_obj: Dict[str, Any], *, item_valido: str='NO_Revisado', comentario: str='') -> List[str]:
    src = site_obj.get('source_csv') or {}
    row = (src.get('row') or {}) if isinstance(src, dict) else {}
    dom = str(site_obj.get('domain') or '').lower().lstrip('www.')
    nombre = _site_name(site_obj, row, dom)
    direccion = (site_obj.get('addresses') or [{}])[0].get('value','') if site_obj.get('addresses') else row.get('Dirección', row.get('Address',''))
    telefono = _first_phone(site_obj, row)
    web = _site_web(site_obj, row, dom)
    ciudad = _infer_city(site_obj, row)
    venue = _infer_venue(src, row)
    puntaje = site_obj.get('band',{}).get('score') or site_obj.get('band_score') or 0
    emails_web_list = _extract_emails_values(site_obj.get('emails') or [])
    emails_web = ', '.join(emails_web_list)
    # externos: try to use exporter split if available
    emails_ext = ''
    try:
        from exportar_etapa1 import load_busquedas_externas, load_json_or_empty, collect_emails_split_strict  # type: ignore
        busq_ext = load_busquedas_externas('out/busquedas_externas.json')
        v2 = load_json_or_empty('out/etapa1_2_V2_V3.json')
        enriq = load_json_or_empty('out/enriquecidos.json')
        enrv3 = load_json_or_empty('out/enriquecidov3.json')
        emw, emx = collect_emails_split_strict(site_obj, busq_ext, v2, enriq, enrv3)
        if emw: emails_web = emw
        emails_ext = emx or ''
    except Exception:
        pass
    cant = len([e for e in [s.strip() for s in (emails_web or '').split(',')] if e])
    socials = _socials_order(site_obj)
    ts = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00','Z')
    short_desc = str(site_obj.get('short_descrip') or site_obj.get('short_desc') or '')
    return [
        item_valido or 'NO_Revisado', comentario or '',
        _resolve_id(site_obj, row), nombre, direccion, telefono, web, ciudad,
        '', venue, puntaje, emails_web, emails_ext, str(cant),
        *socials, ts, short_desc
    ]


def _format_tab3_row(site_obj: Dict[str, Any], *, item_valido: str='NO', comentario: str='', motivo: str='') -> List[str]:
    src = site_obj.get('source_csv') or {}
    row = (src.get('row') or {}) if isinstance(src, dict) else {}
    dom = str(site_obj.get('domain') or '').lower().lstrip('www.')
    nombre = _site_name(site_obj, row, dom)
    direccion = (site_obj.get('addresses') or [{}])[0].get('value','') if site_obj.get('addresses') else row.get('Dirección', row.get('Address',''))
    telefono = _first_phone(site_obj, row)
    web = _site_web(site_obj, row, dom)
    ciudad = _infer_city(site_obj, row)
    venue = _infer_venue(src, row)
    puntaje = site_obj.get('band',{}).get('score') or site_obj.get('band_score') or 0
    emails_web_list = _extract_emails_values(site_obj.get('emails') or [])
    emails_web = ', '.join(emails_web_list)
    emails_ext = ''
    try:
        from exportar_etapa1 import load_busquedas_externas, load_json_or_empty, collect_emails_split_strict  # type: ignore
        busq_ext = load_busquedas_externas('out/busquedas_externas.json')
        v2 = load_json_or_empty('out/etapa1_2_V2_V3.json')
        enriq = load_json_or_empty('out/enriquecidos.json')
        enrv3 = load_json_or_empty('out/enriquecidov3.json')
        emw, emx = collect_emails_split_strict(site_obj, busq_ext, v2, enriq, enrv3)
        if emw: emails_web = emw
        emails_ext = emx or ''
    except Exception:
        pass
    cant = len([e for e in [s.strip() for s in (emails_web or '').split(',')] if e])
    socials = _socials_order(site_obj)
    ts = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00','Z')
    return [
        item_valido or 'NO', comentario or '', motivo or '',
        _resolve_id(site_obj, row), nombre, direccion, telefono, web, ciudad,
        '', venue, puntaje, emails_web, emails_ext, str(cant),
        *socials, ts
    ]


def _load_comentarios_index() -> None:
    """Load comentarios.json into an index keyed by (rubro, 'id'|'dominio', value)."""
    global _coment_idx
    if _coment_idx:
        return
    comentarios = []
    try:
        comentarios = json.load(open('comentarios.json','r',encoding='utf-8'))
    except Exception:
        comentarios = []
    idx = {}
    for ent in (comentarios or []):
        rub = (str(ent.get('rubro') or '')).strip().lower()
        if ent.get('id'):
            idx[(rub,'id',str(ent.get('id')).strip())] = ent
        elif ent.get('dominio'):
            idx[(rub,'dominio',str(ent.get('dominio')).strip().lower())] = ent
    _coment_idx = idx

def _coment_lookup(site_obj: Dict[str,Any], dom: str) -> Tuple[str,str]:
    """Return (item_valido, comentario) from comentarios index when available."""
    try:
        _load_comentarios_index()
        src = site_obj.get('source_csv') or {}
        row = (src.get('row') or {}) if isinstance(src, dict) else {}
        rub = (src.get('rubro') or '')
        key = ((rubro := (rub or '').strip().lower()), 'id', str(row.get('ID') or '').strip())
        ent = _coment_idx.get(key)
        if not ent:
            ent = _coment_idx.get(((rubro := (rub or '').strip().lower()), 'dominio', dom.lower()))
        if ent:
            return (str(ent.get('item_valido') or 'NO_Revisado'), str(ent.get('comentario') or ''))
    except Exception:
        pass
    return ('NO_Revisado','')

def _motivo_descarte(site_obj: Dict[str,Any], discarded: bool, reason: str, eff: float) -> str:
    motivos: List[str] = []
    if not bool(site_obj.get('florida_ok')):
        motivos.append('fuera_de_florida')
    try:
        thr = float(os.getenv('BAND_THRESHOLD') or 10)
    except Exception:
        thr = 10.0
    if eff < thr:
        motivos.append('puntaje_bajo')
    if discarded:
        motivos.append('filtro_config_map')
    if reason:
        motivos.append(reason)
    return '; '.join(dict.fromkeys([m for m in motivos if m]))


# ---------- Public API ----------

def push_site(site_obj: Dict[str, Any]) -> None:
    """Realtime append of a single processed site to TAB1 or TAB3 depending on filters.
    Dedupe by domain against each tab's present domains.
    """
    try:
        if not DEFAULT_SHEET_ID:
            return
        if _exports_disabled():
            return  # globally disabled
        _ensure_initialized()
        ws1 = _ws_cache.get(TAB1); ws3 = _ws_cache.get(TAB3)
        if not (ws1 and ws3):
            return
        # dedupe by 'Sitio Web' domain or by site_obj.domain
        src = site_obj.get('source_csv') or {}
        row = (src.get('row') or {}) if isinstance(src, dict) else {}
        web = row.get('Sitio Web', row.get('Website', row.get('URL','')))
        dom = _domain_from_url(web) if web else (str(site_obj.get('domain') or '').lower().lstrip('www.'))
        if not dom:
            return
        present1 = _present_domains.get(TAB1) or set()
        present3 = _present_domains.get(TAB3) or set()
        # Enriquecer con CSV si faltan ID/Venue
        _attach_source_from_csv_if_missing(site_obj)
        # comentarios y item_valido si existen
        item_val, comm = _coment_lookup(site_obj, dom)
        # Always append to TAB3 (all processed), deduping within TAB3 only
        try:
            if dom not in present3:
                motivo = _motivo_descarte(site_obj, *_evaluate_filters(site_obj)[:2], _evaluate_filters(site_obj)[2])
                vals3 = _format_tab3_row(site_obj, item_valido=item_val if item_val in ('SI','NO') else 'NO', comentario=comm, motivo=motivo)
                ws3.append_row(vals3, value_input_option='RAW')
                try:
                    _append_local_backup_rows(TAB3, [vals3])
                except Exception:
                    pass
                present3.add(dom); _present_domains[TAB3] = present3
        except Exception:
            pass
        # Evaluate filters to decide TAB1 membership
        discarded, reason, eff = _evaluate_filters(site_obj)
        florida_ok = bool(site_obj.get('florida_ok'))
        # Exclusiones USA/NIC: si está en la hoja de exclusiones, no va a TAB1
        excluded_usa = _is_excluded_usa(site_obj)
        if excluded_usa:
            item_val = 'NO'
        to_tab1 = (not discarded) and florida_ok and (eff >= float(os.getenv('BAND_THRESHOLD') or 10)) and (item_val != 'NO') and (not excluded_usa)
        if to_tab1:
            try:
                if dom not in present1:
                    vals1 = _format_tab1_row(site_obj, item_valido=item_val, comentario=comm)
                    ws1.append_row(vals1, value_input_option='RAW')
                    try:
                        _append_local_backup_rows(TAB1, [vals1])
                    except Exception:
                        pass
                    present1.add(dom); _present_domains[TAB1] = present1
            except Exception:
                pass
        # Remove from TAB2 if present
        try:
            ws2 = _ws_cache.get(TAB2)
            if ws2:
                vals2 = ws2.get_all_values() or []
                if vals2:
                    hdr = vals2[0]
                    idx_web = hdr.index('Sitio Web') if 'Sitio Web' in hdr else -1
                    if idx_web >= 0:
                        # find row by matching domain in Sitio Web
                        for i in range(len(vals2)-1, 0, -1):  # bottom-up
                            v = (vals2[i][idx_web] if idx_web < len(vals2[i]) else '').strip()
                            if v and _domain_from_url(v) == dom:
                                ws2.delete_rows(i+1)
                                break
        except Exception:
            pass
    except Exception:
        # best-effort: do not break caller
        pass


def rebuild_unprocessed_from_csv(csv_dir: str = 'csv') -> int:
    """Rebuild TAB2 with all rows across csv/*.csv, deduped by domain (from Sitio Web) or by ID fallback.
    Returns number of rows uploaded (excluding header).
    """
    _ensure_initialized()
    ws2 = _ws_cache.get(TAB2)
    if not ws2:
        return 0
    # Exclusion set from TAB4 (Filtrados) by domain
    doms_filtered: Set[str] = set()
    try:
        ws4 = _ws_cache.get(TAB4)
        if ws4:
            vals4 = ws4.get_all_values() or []
            if vals4 and len(vals4) > 1:
                hdr4 = vals4[0]
                idx4 = hdr4.index('Sitio Web') if 'Sitio Web' in hdr4 else -1
                if idx4 >= 0:
                    for i in range(1, len(vals4)):
                        web4 = (vals4[i][idx4] if idx4 < len(vals4[i]) else '').strip()
                        if web4:
                            d4 = _domain_from_url(web4)
                            if d4:
                                doms_filtered.add(d4)
    except Exception:
        pass
    from pathlib import Path as _P
    import csv as _csv
    rows_out: List[List[str]] = [TAB2_HEADER]
    seen_dom: Set[str] = set(); seen_id: Set[str] = set()
    for p in sorted(_P(csv_dir).glob('*.csv')):
        try:
            with p.open('r', encoding='utf-8', errors='ignore') as f:
                reader = _csv.DictReader(f)
                for r in reader:
                    _id = str(r.get('ID') or r.get('Id') or r.get('id') or '').strip()
                    web = r.get('Sitio Web') or r.get('Website') or r.get('URL') or ''
                    dom = _domain_from_url(web) if web else ''
                    if dom and dom in doms_filtered:
                        # Excluir dominios presentes en TAB4
                        continue
                    key_ok = False
                    if dom:
                        if dom in seen_dom: continue
                        seen_dom.add(dom); key_ok = True
                    elif _id:
                        if _id in seen_id: continue
                        seen_id.add(_id); key_ok = True
                    if not key_ok:
                        # skip totally empty keys
                        continue
                    venue = p.stem
                    rows_out.append([
                        _id, venue, r.get('Nombre', r.get('Name','')), r.get('Dirección', r.get('Address','')),
                        r.get('Teléfono', r.get('Telefono', r.get('Phone', r.get('TEL','')))),
                        web, r.get('Ciudad', r.get('City','')), r.get('GoogleMapsURL','')
                    ])
        except Exception:
            continue
    # Rewrite entire sheet (header + body)
    ws2.clear()
    ws2.update(rows_out)
    return max(0, len(rows_out)-1)


def sync_processed_from_etapa1(etapa1_path: str = 'out/etapa1_v1.json') -> Tuple[int, int]:
    """Scan etapa1_v1.json and append missing to TAB1 and TAB3 according to filters.
    Returns (#added_tab1, #added_tab3).
    """
    _ensure_initialized()
    if _exports_disabled():
        return (0, 0)
    ws1 = _ws_cache.get(TAB1); ws3 = _ws_cache.get(TAB3)
    if not (ws1 and ws3):
        return (0, 0)
    try:
        data = json.load(open(etapa1_path, 'r', encoding='utf-8'))
        sites = data.get('sites', []) if isinstance(data, dict) else []
    except Exception:
        return (0, 0)
    # Load comentarios for item_valido/comentario
    _load_comentarios_index()
    def _coment_key(rubro: str, _id: str, dom: str) -> Tuple[str, str, str]:
        rub = (rubro or '').strip().lower(); _id2 = (_id or '').strip(); dom2 = (dom or '').strip().lower()
        if _id2: return (rub, 'id', _id2)
        return (rub, 'dominio', dom2)
    # Seeds
    present1 = _present_domains.get(TAB1) or set(); present3 = _present_domains.get(TAB3) or set()
    add1 = 0; add3 = 0
    rows_batch_1: List[List[str]] = []
    rows_batch_3: List[List[str]] = []
    moved_domains: List[str] = []
    for s in sites:
        src = s.get('source_csv') or {}
        row = (src.get('row') or {}) if isinstance(src, dict) else {}
        web = row.get('Sitio Web', row.get('Website', row.get('URL','')))
        dom = _domain_from_url(web) if web else (str(s.get('domain') or '').lower().lstrip('www.'))
        if not dom:
            continue
        # Enriquecer con CSV si faltan ID/Venue
        _attach_source_from_csv_if_missing(s)
        # Do not skip globally; decide per tab to ensure TAB3 has ALL processed
        rub = (src.get('rubro') or '') if isinstance(src, dict) else ''
        key = _coment_key((rub or '').lower(), str(row.get('ID') or ''), dom)
        ent = _coment_idx.get(key) or {}
        # Default for new items unless explicitly set to NO or SI elsewhere
        item_val = str(ent.get('item_valido') or 'NO_Revisado')
        # Override: mark SI if present in mark-SI sheet (domain/emails)
        try:
            if _is_marked_yes(s):
                item_val = 'SI'
        except Exception:
            pass
        comm = str(ent.get('comentario') or '')
        # Decide destination
        discarded, reason, eff = _evaluate_filters(s)
        florida_ok = bool(s.get('florida_ok'))
        # Always include in TAB3 (all processed), dedupe within TAB3
        if dom not in present3:
            motivo = _motivo_descarte(s, discarded, reason, eff)
            vals3 = _format_tab3_row(s, item_valido=item_val if item_val in ('SI','NO') else 'NO', comentario=comm, motivo=motivo)
            rows_batch_3.append(vals3); present3.add(dom); add3 += 1
            moved_domains.append(dom)
        # Include in TAB1 if passes filters and not marked NO, dedupe within TAB1
        excluded_usa = _is_excluded_usa(s)
        # If excluded by USA list, mark as NO explicitly
        if excluded_usa:
            item_val = 'NO'
        if (not discarded) and florida_ok and (eff >= float(os.getenv('BAND_THRESHOLD') or 10)) and (item_val != 'NO') and (not excluded_usa):
            if dom not in present1:
                vals1 = _format_tab1_row(s, item_valido=item_val, comentario=comm)
                rows_batch_1.append(vals1); present1.add(dom); add1 += 1
    # Append in batches to reduce API calls (avoid 429)
    def _append_rows(ws, rows: List[List[str]], chunk: int = 400):
        for i in range(0, len(rows), chunk):
            part = rows[i:i+chunk]
            try:
                ws.append_rows(part, value_input_option='RAW')
            except Exception:
                # fallback one-by-one if batch fails
                for r in part:
                    try:
                        ws.append_row(r, value_input_option='RAW')
                    except Exception:
                        pass
            time.sleep(0.2)  # small pacing between chunks
    if rows_batch_1:
        _append_rows(ws1, rows_batch_1)
        _append_local_backup_rows(TAB1, rows_batch_1)
    if rows_batch_3:
        _append_rows(ws3, rows_batch_3)
        _append_local_backup_rows(TAB3, rows_batch_3)
    # Remove moved domains from TAB2
    try:
        ws2 = _ws_cache.get(TAB2)
        if ws2 and moved_domains:
            vals2 = ws2.get_all_values() or []
            if vals2:
                hdr = vals2[0]; idx_web = hdr.index('Sitio Web') if 'Sitio Web' in hdr else -1
                if idx_web >= 0:
                    to_delete = []
                    domset = set(moved_domains)
                    for i in range(1, len(vals2)):
                        v = (vals2[i][idx_web] if idx_web < len(vals2[i]) else '').strip()
                        if v and _domain_from_url(v) in domset:
                            to_delete.append(i+1)
                    for ridx in sorted(to_delete, reverse=True):
                        try:
                            ws2.delete_rows(ridx)
                        except Exception:
                            pass
    except Exception:
        pass
    _present_domains[TAB1] = present1; _present_domains[TAB3] = present3
    return (add1, add3)


# ---------- DOM_Name CSV filtering and push to TAB4 ----------

def _load_config_map() -> Dict[str, Any]:
    path = os.getenv('CONFIG_MAP_PATH', os.path.join('configuraciones','config_map.json'))
    try:
        return json.load(open(path,'r',encoding='utf-8'))
    except Exception:
        return {}

def _normalize_rubro_key(rubro: Optional[str]) -> str:
    rl = (rubro or '').strip().lower()
    if rl.startswith('rubro_'):
        rl = rl[6:]
    return rl

def rebuild_filtered_unprocessed_from_csv(csv_dir: str = 'csv') -> int:
    """Reconstruye TAB4 ('Filtrados_sin_procesar_bot') a partir de csv/*.csv aplicando reglas DOM_Name.
    - Aplica discard_keywords de defaults_Dom_Name y del rubro.dom_name sobre Nombre y dominio.
    - Sube filas con el formato requerido agregando la columna Filtro_hit con la palabra que disparó el filtro.
    Devuelve la cantidad de filas subidas (excluyendo header).
    """
    _ensure_initialized()
    ws4 = _ws_cache.get(TAB4)
    if not ws4:
        return 0
    cfg = _load_config_map() or {}
    defaults_dom = ((cfg.get('defaults_Dom_Name') or {}) if isinstance(cfg, dict) else {})
    d_discard = [str(x).strip().lower() for x in (defaults_dom.get('discard_keywords') or [])]
    from pathlib import Path as _P
    import csv as _csv
    rows_out: List[List[str]] = [TAB4_HEADER]
    seen_keys: Set[str] = set()
    for p in sorted(_P(csv_dir).glob('*.csv')):
        venue = p.stem
        rkey = _normalize_rubro_key(venue)
        rub = (cfg.get('rubros') or {}).get(rkey, {}) if isinstance(cfg.get('rubros'), dict) else {}
        dom_name = (rub.get('dom_name') or {}) if isinstance(rub, dict) else {}
        r_discard = [str(x).strip().lower() for x in (dom_name.get('discard_keywords') or [])]
        discard_kws = [k for k in (d_discard + r_discard) if k]
        if not discard_kws:
            continue
        try:
            with p.open('r', encoding='utf-8', errors='ignore') as f:
                reader = _csv.DictReader(f)
                for r in reader:
                    name = str(r.get('Nombre', r.get('Name','')) or '').strip().lower()
                    web = r.get('Sitio Web') or r.get('Website') or r.get('URL') or ''
                    dom = _domain_from_url(web) if web else ''
                    dom_low = dom.lower()
                    addr_low = str(r.get('Dirección', r.get('Address','')) or '').strip().lower()
                    hit_kw = None
                    for kw in discard_kws:
                        if kw and ((kw in name) or (kw in dom_low) or (kw in addr_low)):
                            hit_kw = kw
                            break
                    if not hit_kw:
                        continue
                    # dedupe key by domain or ID
                    _id = str(r.get('ID') or r.get('Id') or r.get('id') or '').strip()
                    k = dom_low or (_id and f"id:{_id}") or None
                    if not k or k in seen_keys:
                        continue
                    seen_keys.add(k)
                    rows_out.append([
                        _id,
                        venue,
                        r.get('Nombre', r.get('Name','')) or '',
                        r.get('Dirección', r.get('Address','')) or '',
                        r.get('Teléfono', r.get('Telefono', r.get('Phone', r.get('TEL','')))) or '',
                        web or '',
                        r.get('Ciudad', r.get('City','')) or '',
                        r.get('GoogleMapsURL','') or '',
                        hit_kw or ''
                    ])
        except Exception:
            continue
    # Rewrite entire tab
    try:
        ws4.clear()
    except Exception:
        pass
    if len(rows_out) > 1:
        try:
            ws4.update(rows_out)
        except Exception:
            # fallback append groups
            for i in range(0, len(rows_out), 400):
                part = rows_out[i:i+400]
                try:
                    ws4.append_rows(part, value_input_option='RAW')
                except Exception:
                    for r in part:
                        try:
                            ws4.append_row(r, value_input_option='RAW')
                        except Exception:
                            pass
    # Remove any matching domains from TAB2 (sin_procesar_bot)
    try:
        ws2 = _ws_cache.get(TAB2)
        if ws2:
            vals4 = ws4.get_all_values() or []
            vals2 = ws2.get_all_values() or []
            if vals4 and vals2 and len(vals4) > 1 and len(vals2) > 1:
                hdr4 = vals4[0]; idx4 = hdr4.index('Sitio Web') if 'Sitio Web' in hdr4 else -1
                hdr2 = vals2[0]; idx2 = hdr2.index('Sitio Web') if 'Sitio Web' in hdr2 else -1
                if idx4 >= 0 and idx2 >= 0:
                    doms4: Set[str] = set()
                    for i in range(1, len(vals4)):
                        web = (vals4[i][idx4] if idx4 < len(vals4[i]) else '').strip()
                        if web:
                            d = _domain_from_url(web)
                            if d:
                                doms4.add(d)
                    to_delete = []
                    for i in range(1, len(vals2)):
                        web2 = (vals2[i][idx2] if idx2 < len(vals2[i]) else '').strip()
                        if web2 and _domain_from_url(web2) in doms4:
                            to_delete.append(i+1)  # 1-based
                    for ridx in sorted(to_delete, reverse=True):
                        try:
                            ws2.delete_rows(ridx)
                        except Exception:
                            pass
    except Exception:
        pass
    return max(0, len(rows_out)-1)


def apply_dom_name_filter_on_tab1_move_to_tab3() -> int:
    """Scan TAB1 (procesados_bot) and move rows to TAB3 if DOM_Name discard hits
    on Nombre or domain according to config_map. Sets motivo_descarte to
    "Filtro: <keyword>" and deletes from TAB1. Returns number of moved rows.
    """
    _ensure_initialized()
    ws1 = _ws_cache.get(TAB1); ws3 = _ws_cache.get(TAB3)
    if not (ws1 and ws3):
        return 0
    cfg = _load_config_map() or {}
    defaults_dom = ((cfg.get('defaults_Dom_Name') or {}) if isinstance(cfg, dict) else {})
    base_discard = [str(x).strip().lower() for x in (defaults_dom.get('discard_keywords') or [])]
    try:
        vals1 = ws1.get_all_values() or []
        if not vals1 or len(vals1) < 2:
            return 0
        hdr1 = vals1[0]
        # Robust column resolution to handle encoding variants (Dirección/Direccion/Direccin) and alternate labels
        def _idx_exact(col: str) -> int:
            try:
                return hdr1.index(col)
            except ValueError:
                return -1
        def _idx_any(candidates: list) -> int:
            for c in candidates:
                try:
                    return hdr1.index(c)
                except ValueError:
                    continue
            return -1
        idx_item = 0  # Item_Valido
        idx_comm = 1  # comentario
        idx_id   = _idx_exact('ID')
        idx_nom  = _idx_exact('Nombre')
        # Try multiple header variants for Address
        idx_dir  = _idx_any(['Dirección','Direccion','Direccin','Address','Direccion_Completa'])
        idx_web  = _idx_any(['Sitio Web','WEB','Website'])
        idx_ven  = _idx_exact('Venue')
        if idx_web < 0:
            return 0
        moved = 0
        # bottom-up delete safety
        for i in range(len(vals1)-1, 0, -1):
            row = vals1[i]
            web = (row[idx_web] if idx_web < len(row) else '').strip()
            dom = _domain_from_url(web) if web else ''
            name_low = str(row[idx_nom] if (idx_nom >= 0 and idx_nom < len(row)) else '').strip().lower()
            addr_low = str(row[idx_dir] if (idx_dir >= 0 and idx_dir < len(row)) else '').strip().lower()
            venue = str(row[idx_ven] if (idx_ven >= 0 and idx_ven < len(row)) else '').strip()
            rkey = _normalize_rubro_key(venue)
            rub = (cfg.get('rubros') or {}).get(rkey, {}) if isinstance(cfg.get('rubros'), dict) else {}
            dom_name = (rub.get('dom_name') or {}) if isinstance(rub, dict) else {}
            r_discard = [str(x).strip().lower() for x in (dom_name.get('discard_keywords') or [])]
            discard_kws = [k for k in (base_discard + r_discard) if k]
            if not discard_kws:
                continue
            hit_kw = None
            dom_low = (dom or '').lower()
            for kw in discard_kws:
                if kw and ((kw in name_low) or (kw in dom_low) or (kw in addr_low)):
                    hit_kw = kw
                    break
            if not hit_kw:
                continue
            comentario = (row[idx_comm] if idx_comm < len(row) else '').strip()
            motivo = f"Filtro: {hit_kw}"
            base = row[:] + [''] * max(0, len(TAB1_HEADER) - len(row))
            new_row = ['NO', comentario, motivo] + base[2:]
            try:
                ws3.append_row(new_row, value_input_option='RAW')
                ws1.delete_rows(i+1)
                # keep present domains set in sync
                if dom:
                    _present_domains.get(TAB1, set()).discard(dom)
                    _present_domains.setdefault(TAB3, set()).add(dom)
                moved += 1
            except Exception:
                continue
        return moved
    except Exception:
        return 0


def force_move_by_keyword_on_tab1(keywords: List[str]) -> int:
    """Force-move rows in TAB1 to TAB3 when any keyword appears in Nombre or domain.
    Sets motivo_descarte = "Filtro: <kw> (force)". Returns number of moved rows.
    """
    _ensure_initialized()
    ws1 = _ws_cache.get(TAB1); ws3 = _ws_cache.get(TAB3)
    if not (ws1 and ws3):
        return 0
    kws = [str(k).strip().lower() for k in (keywords or []) if str(k).strip()]
    if not kws:
        return 0
    try:
        vals1 = ws1.get_all_values() or []
        if not vals1 or len(vals1) < 2:
            return 0
        hdr1 = vals1[0]
        def _idx(col: str) -> int:
            try:
                return hdr1.index(col)
            except ValueError:
                return -1
        idx_comm = 1
        idx_nom  = _idx('Nombre')
        idx_web  = _idx('Sitio Web')
        if idx_web < 0:
            return 0
        moved = 0
        for i in range(len(vals1)-1, 0, -1):
            row = vals1[i]
            nom_low = str(row[idx_nom] if (idx_nom >= 0 and idx_nom < len(row)) else '').strip().lower()
            web = (row[idx_web] if idx_web < len(row) else '').strip()
            dom_low = _domain_from_url(web).lower() if web else ''
            hit = None
            for kw in kws:
                if kw and (kw in nom_low or kw in dom_low):
                    hit = kw; break
            if not hit:
                continue
            comentario = (row[idx_comm] if idx_comm < len(row) else '').strip()
            motivo = f"Filtro: {hit} (force)"
            base = row[:] + [''] * max(0, len(TAB1_HEADER) - len(row))
            new_row = ['NO', comentario, motivo] + base[2:]
            try:
                ws3.append_row(new_row, value_input_option='RAW')
                ws1.delete_rows(i+1)
                moved += 1
            except Exception:
                continue
        return moved
    except Exception:
        return 0


def apply_sheet_review_moves() -> int:
    """Move rows in TAB1 with Item_Valido == 'NO' to TAB3, copying comentario to motivo_descarte,
    and delete them from TAB1. Returns number of moved rows."""
    _ensure_initialized()
    ws1 = _ws_cache.get(TAB1); ws3 = _ws_cache.get(TAB3)
    if not (ws1 and ws3):
        return 0
    try:
        vals1 = ws1.get_all_values() or []
        if not vals1 or len(vals1) < 2:
            return 0
        hdr1 = vals1[0]
        idx_item = 0  # 'Item_Valido'
        idx_comm = 1  # 'comentario'
        idx_web = hdr1.index('Sitio Web') if 'Sitio Web' in hdr1 else -1
        moved = 0
        for i in range(len(vals1)-1, 0, -1):  # bottom-up to preserve indices on delete
            row = vals1[i]
            val = (row[idx_item] if idx_item < len(row) else '').strip().upper()
            if val.upper() == 'NO':
                comentario = (row[idx_comm] if idx_comm < len(row) else '').strip()
                motivo = comentario or 'marcado_manual_NO'
                # Build row for TAB3 from existing row
                # TAB3 header: Item_Valido, comentario, motivo_descarte, [rest of TAB1 cols...]
                base = row[:] + [''] * max(0, len(TAB1_HEADER) - len(row))
                new_row = ['NO', comentario, motivo] + base[2:]
                try:
                    ws3.append_row(new_row, value_input_option='RAW')
                    ws1.delete_rows(i+1)
                    # update in-memory sets
                    if idx_web >= 0 and idx_web < len(row):
                        dom = _domain_from_url(row[idx_web])
                        if dom:
                            _present_domains.get(TAB1, set()).discard(dom)
                            _present_domains.setdefault(TAB3, set()).add(dom)
                    moved += 1
                except Exception:
                    continue
        return moved
    except Exception:
        return 0


def sync_from_db_etapa1_sites(db_path: str = 'data.db') -> Tuple[int, int]:
    """Backfill desde la BD (tabla etapa1_sites) para cargar TODOS los procesados en TAB3
    y los que pasan filtros en TAB1. Devuelve (#added_tab1, #added_tab3).
    """
    _ensure_initialized()
    if _exports_disabled():
        return (0, 0)
    ws1 = _ws_cache.get(TAB1); ws3 = _ws_cache.get(TAB3)
    if not (ws1 and ws3):
        return (0, 0)
    try:
        cx = sqlite3.connect(db_path)
    except Exception:
        return (0, 0)
    present1 = _present_domains.get(TAB1) or set(); present3 = _present_domains.get(TAB3) or set()
    add1 = 0; add3 = 0
    rows_batch_1: List[List[str]] = []
    rows_batch_3: List[List[str]] = []
    moved_domains: List[str] = []
    try:
        cur = cx.cursor()
        cur.execute('SELECT domain, raw_json FROM etapa1_sites')
        for domain, raw_json in cur.fetchall():
            dom = (domain or '').strip().lower().lstrip('www.')
            if not dom:
                continue
            try:
                s = json.loads(raw_json or '{}') if raw_json else {}
            except Exception:
                s = {}
            # Ensure site_obj has domain and source
            if 'domain' not in s:
                s['domain'] = dom
            # Enriquecer con CSV si faltan ID/Venue
            _attach_source_from_csv_if_missing(s)
            # comentarios
            item_val, comm = _coment_lookup(s, dom)
            # Override: mark SI if present in mark-SI sheet (domain/emails)
            try:
                if _is_marked_yes(s):
                    item_val = 'SI'
            except Exception:
                pass
            # Filtros
            discarded, reason, eff = _evaluate_filters(s)
            florida_ok = bool(s.get('florida_ok'))
            # Always TAB3 (dedupe in TAB3)
            if dom not in present3:
                motivo = _motivo_descarte(s, discarded, reason, eff)
                v3 = _format_tab3_row(s, item_valido=item_val if item_val in ('SI','NO') else 'NO', comentario=comm, motivo=motivo)
                rows_batch_3.append(v3); present3.add(dom); add3 += 1
                moved_domains.append(dom)
            # TAB1 si pasa filtros y no está marcado NO
            excluded_usa = _is_excluded_usa(s)
            if excluded_usa:
                item_val = 'NO'
            if (not discarded) and florida_ok and (eff >= float(os.getenv('BAND_THRESHOLD') or 10)) and (item_val != 'NO') and (not excluded_usa):
                if dom not in present1:
                    v1 = _format_tab1_row(s, item_valido=item_val, comentario=comm)
                    rows_batch_1.append(v1); present1.add(dom); add1 += 1
    finally:
        try:
            cx.close()
        except Exception:
            pass
    # Append in batches
    def _append_rows(ws, rows: List[List[str]], chunk: int = 400):
        for i in range(0, len(rows), chunk):
            part = rows[i:i+chunk]
            try:
                ws.append_rows(part, value_input_option='RAW')
            except Exception:
                for r in part:
                    try:
                        ws.append_row(r, value_input_option='RAW')
                    except Exception:
                        pass
            time.sleep(0.2)
    if rows_batch_3:
        _append_rows(ws3, rows_batch_3)
        _append_local_backup_rows(TAB3, rows_batch_3)
    if rows_batch_1:
        _append_rows(ws1, rows_batch_1)
        _append_local_backup_rows(TAB1, rows_batch_1)
    # Remove from TAB2
    try:
        ws2 = _ws_cache.get(TAB2)
        if ws2 and moved_domains:
            vals2 = ws2.get_all_values() or []
            if vals2:
                hdr = vals2[0]; idx_web = hdr.index('Sitio Web') if 'Sitio Web' in hdr else -1
                if idx_web >= 0:
                    to_delete = []
                    domset = set(moved_domains)
                    for i in range(1, len(vals2)):
                        v = (vals2[i][idx_web] if idx_web < len(vals2[i]) else '').strip()
                        if v and _domain_from_url(v) in domset:
                            to_delete.append(i+1)
                    for ridx in sorted(to_delete, reverse=True):
                        try:
                            ws2.delete_rows(ridx)
                        except Exception:
                            pass
    except Exception:
        pass
    _present_domains[TAB1] = present1; _present_domains[TAB3] = present3
    return (add1, add3)


# ---------- Maintenance helpers ----------
def reapply_item_valido_validation() -> None:
    """Re-applies the Item_Valido and comentario dropdown validations to TAB1 and TAB3.
    Useful if the sheet existed before or if validations were removed by edits.
    """
    _ensure_initialized()
    try:
        sh = _load_sheet()
        for title in (TAB1, TAB3):
            try:
                ws = _ws_cache.get(title)
                if not ws:
                    ws = sh.worksheet(title)
                if ws:
                    _apply_validations_for_ws(sh, ws, title)
            except Exception:
                pass
    except Exception:
        pass

if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser(description='Sync utilities for Todos_ADT_OTS_Expo')
    p.add_argument('--rebuild-sin-procesar', action='store_true')
    p.add_argument('--sync-procesados', action='store_true')
    p.add_argument('--sync-from-db', action='store_true', help='Backfill TAB3 (and TAB1 if pasa filtros) desde data.db etapa1_sites')
    p.add_argument('--db-path', type=str, default='data.db')
    p.add_argument('--rebuild-filtrados', action='store_true', help="Reconstruye 'Filtrados_sin_procesar_bot' aplicando DOM_Name a csv/*")
    p.add_argument('--backfill-short-descrip', action='store_true', help='Rellena Short_descrip en TAB1 usando cache + IA (solo vacíos)')
    p.add_argument('--apply-domname-to-procesados', action='store_true', help='Mueve de TAB1 a TAB3 si DOM_Name hit; motivo_descarte="Filtro: <kw>"')
    p.add_argument('--force-domname-keywords', type=str, default='', help='Lista separada por comas de keywords para forzar mover de TAB1 a TAB3')
    args = p.parse_args()
    if args.rebuild_sin_procesar:
        n = rebuild_unprocessed_from_csv()
        print(f"TAB2 rebuilt: {n} rows")
    if args.sync_procesados:
        a1, a3 = sync_processed_from_etapa1()
        print(f"TAB1 +{a1} | TAB3 +{a3}")
    if args.sync_from_db:
        a1, a3 = sync_from_db_etapa1_sites(args.db_path)
        print(f"[DB] TAB1 +{a1} | TAB3 +{a3}")
    if args.rebuild_filtrados:
        n4 = rebuild_filtered_unprocessed_from_csv()
        print(f"TAB4 rebuilt: {n4} rows")
    if args.backfill_short_descrip:
        n = backfill_short_descrip_tab1()
        print(f"Short_descrip backfilled: {n} rows")
    if args.apply_domname_to_procesados:
        mv = apply_dom_name_filter_on_tab1_move_to_tab3()
        print(f"TAB1→TAB3 (DOM_Name) moved: {mv}")
    if args.force_domname_keywords:
        kws = [x.strip() for x in args.force_domname_keywords.split(',') if x.strip()]
        mv2 = force_move_by_keyword_on_tab1(kws)
        print(f"TAB1→TAB3 (force keywords) moved: {mv2}")

